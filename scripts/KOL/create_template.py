#!/usr/bin/env python

from openpyxl.utils import get_column_letter
import openpyxl

# sample_file = argv[1]
# template_file = argv[2]

sample_file = 'quid_KOL_news_output_v2.xlsx'
template_file = 'template.xlsx'

raw_ceil = 2
magic_ceil = 4

book = openpyxl.load_workbook(sample_file)

raw_sheet = book.get_sheet_by_name("Raw data")
output_sheet = book.get_sheet_by_name("Output")

print(raw_sheet.max_row)
print(output_sheet.max_row)

for row in raw_sheet['A2:AY{}'.format(raw_sheet.max_row)]:
    for cell in row:
        cell.value = None

for row in output_sheet['A3:A{}'.format(output_sheet.max_row)]:
    for cell in row:
        cell.value = None
book.save(template_file)
