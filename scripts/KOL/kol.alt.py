#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import csv
import re
import openpyxl
from sys import argv, maxsize

# input_file = 'Sample News CSV.csv'
# output_file = 'NewsKol.xlsx'
# template_file = 'template.xlsx'

input_file = argv[1]
output_file = argv[2]
template_file = argv[3]

def copy_formulae(ws, r_index, t_index):
    row = []
    r,t= str(r_index),str(t_index)
    for x in (list(ws.rows)[r_index-1]):
        if x.value is not None:
            row.append(x.value.replace(r,t))
        else:
            row.append(x.value)
    return row

# for OverflowError
decrement = True
while decrement:
    decrement = False
    try:
        csv.field_size_limit(maxsize)
    except OverflowError:
        maxsize = int(maxsize/10)
        decrement = True

book = openpyxl.load_workbook(template_file)

raw_sheet = book.get_sheet_by_name("Raw data")
output_sheet = book.get_sheet_by_name("Output")
raw_title = []
people_name = set()

for row in raw_sheet.iter_rows():
    for cell in row:
        raw_title.append(cell.value)

with open(input_file) as csvFile:
    reader = csv.DictReader(csvFile)
    for r,row in enumerate(reader,start=2):
        result = re.sub(r'\([^)]*\)', '',row['People (Any Mention)'])
        people_name |= set([people.strip() for people in result.split(';')])
        for c,title in enumerate(raw_title,start=1):
            raw_sheet.cell(row=r,column=c).value=row[title]

try:
    people_name.remove('')
    people_name.remove(' ')
    people_name.remove('People (Any Mention)')
except:
    pass

for r,row in enumerate(output_sheet['A24:O{}'.format(str(len(people_name)+2))],start=3):
    for cell, new in zip(row, copy_formulae(output_sheet,20,r)):
        cell.value = new

for src, dst in zip(people_name, output_sheet['A3':'A{}'.format(str(3+len(people_name)))]):
    dst[0].value = src

book.save(output_file)
