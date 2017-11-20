#!/usr/bin/env python

# NOTE:  gave a div0 error with AVERAGEIF for the given sample input

import os
import re
import csv
import openpyxl
from sys import argv, maxsize

formula = [
'=COUNTIF(INDEX(\'Raw data\'!$A:$BG, 0, MATCH("People (Any Mention)", \'Raw data\'!$A$1:$BG$1, 0)), "*" & A{0} & "*")',\
'=AVERAGEIF(INDEX(\'Raw data\'!$A:$BG, 0, MATCH("People (Any Mention)", \'Raw data\'!$A$1:$BG$1, 0)),"*" & A{0} & "*",INDEX(\'Raw data\'!$A:$BG, 0, MATCH("Betweenness Centrality", \'Raw data\'!$A$1:$BG$1, 0)))',\
'=AVERAGEIF(INDEX(\'Raw data\'!$A:$BG, 0, MATCH("People (Any Mention)", \'Raw data\'!$A$1:$BG$1, 0)),"*" & A{0} & "*",INDEX(\'Raw data\'!$A:$BG, 0, MATCH("Inter-Cluster Connectivity", \'Raw data\'!$A$1:$BG$1, 0)))',\
'=SUMIF(INDEX(\'Raw data\'!$A:$BG, 0, MATCH("People (Any Mention)", \'Raw data\'!$A$1:$BG$1, 0)),"*" & A{0} & "*",INDEX(\'Raw data\'!$A:$BG, 0, MATCH("Social Engagement", \'Raw data\'!$A$1:$BG$1, 0)))',\
'=SUMIF(INDEX(\'Raw data\'!$A:$BG, 0, MATCH("People (Any Mention)", \'Raw data\'!$A$1:$BG$1, 0)),"*" & A{0} & "*",INDEX(\'Raw data\'!$A:$BG, 0, MATCH("Published Count", \'Raw data\'!$A$1:$BG$1, 0)))',\
'=IFERROR(AVERAGEIF(INDEX(\'Raw data\'!$A:$BG, 0, MATCH("People (Any Mention)", \'Raw data\'!$A$1:$BG$1, 0)),"*" & A{0} & "*",INDEX(\'Raw data\'!$A:$BG, 0, MATCH("Source Rank", \'Raw data\'!$A$1:$BG$1, 0))), 0)',\
'=PERCENTRANK(B:B,B{0})',\
'=PERCENTRANK(C:C,C{0})',\
'=PERCENTRANK(D:D,D{0})',\
'=PERCENTRANK(E:E,E{0})',\
'=PERCENTRANK(F:F,F{0})',\
'=1 - PERCENTRANK(G:G,G{0})',\
'=(H{0}*weight1) + (I{0}*weight2)+(J{0}*weight3)+(K{0}*weight4)+ (L{0}*weight5)+(M{0}*weight6)',\
'=RANK(N{0},N:N)'
]

# input_file = 'Sample News CSV.csv'
# output_file = 'NewsKol.xlsx'
# template_file = 'template.xlsx'

input_file = argv[1]
output_file = argv[2]
template_file = argv[3]

# for OverflowError
decrement = True
while decrement:
    decrement = False
    try:
        csv.field_size_limit(maxsize)
    except OverflowError:
        maxsize = int(maxsize/10)
        decrement = True

book = openpyxl.load_workbook(template_file)

raw_sheet = book.get_sheet_by_name("Raw data")
output_sheet = book.get_sheet_by_name("Output")
raw_title = []
people_name = set()

for row in raw_sheet.iter_rows():
    for cell in row:
        raw_title.append(cell.value)

with open(input_file) as csvFile:
    reader = csv.DictReader(csvFile)
    for r,row in enumerate(reader,start=2):
        result = re.sub(r'\([^)]*\)', '',row['People (Any Mention)'])
        people_name |= set([people.replace(')','').strip() for people in result.split(';')])
        for c,title in enumerate(raw_title,start=1):
            raw_sheet.cell(row=r,column=c).value=row[title]

try:
    people_name.remove('')
    people_name.remove(' ')
    people_name.remove('People (Any Mention)')
except:
    pass

for r in range(3,len(people_name) + 3):
    for f in range(len(formula)):
        # output_sheet.cell(row=r,column=(f+1)).value=formula[f]
        output_sheet.cell(row=r,column=(f+2)).set_explicit_value(value=formula[f].format(r),data_type=cell.TYPE_FORMULA)

for src, dst in zip(people_name, output_sheet['A3':'A{}'.format(str(3+len(people_name)))]):
    dst[0].value = src

book.save(output_file)
