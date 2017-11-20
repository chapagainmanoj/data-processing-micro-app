#!/usr/bin/env python
# -*- coding: utf-8 -*-

import re
import csv
import openpyxl
from sys import argv, maxsize
from collections import Counter
from openpyxl.utils.exceptions import IllegalCharacterError

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

# input_file = 'Drones - 3034 Companies.csv'
# output_file = 'RegenWiz.xlsx'
# template_file = 'template.xlsx'

input_file = argv[1]
output_file = argv[2]
template_file = argv[3]

template_floor = 6460
Y_formula = """=IF($C{0}="x",$A{1}&", ","")"""
Z_formula = "=CONCATENATE(Z{0},Y{1})"
AA_formula = """=IF($D{0}="x",$A{1}&", ","")"""
AB_formula = "=CONCATENATE(AB{0},AA{1})"

# for OverflowError
decrement = True
while decrement:
    decrement = False
    try:
        csv.field_size_limit(maxsize)
    except OverflowError:
        maxsize = int(maxsize/10)
        decrement = True

book = openpyxl.load_workbook(template_file)
raw_sheet = book.get_sheet_by_name("Raw Data")
magic_sheet = book.get_sheet_by_name("Magic")

raw_title = ['Node ID', 'Keywords']
magic_title = ['Keyword', 'Count - Keyword','Blacklist','Boost']

with open(input_file,encoding='utf-8') as csvFile:
    raw_ceil = 2
    magic_ceil = 5
    counts = Counter([])
    reader = csv.DictReader(csvFile)

    for row in reader:
        words = list(map(str.strip,row[raw_title[1]].split(';')))
        nodeid = row[raw_title[0]]
        counts.update(words)
        for r, word in enumerate(words):
            raw_sheet.cell(row=raw_ceil,column=1).value = nodeid
            try:
                raw_sheet.cell(row=raw_ceil,column=2).value = word
            except IllegalCharacterError:
                word = re.sub(ILLEGAL_CHARACTERS_RE,'',word)
                raw_sheet.cell(row=raw_ceil,column=2).value = word
            raw_ceil+=1

    for key,freq in counts.most_common():
        try:
            magic_sheet.cell(row=magic_ceil,column=1).value = key
        except IllegalCharacterError:
            key = ILLEGAL_CHARACTERS_RE.sub('',key)
            raw_sheet.cell(row=raw_ceil,column=1).value = key
        magic_sheet.cell(row=magic_ceil,column=2).value = freq
        magic_ceil+=1
        # manual write
    magic_sheet['B1'].value = '=Z{0}&" "&G1'.format(str(magic_ceil))
    magic_sheet['B2'].value = '=AB{0}&" "&G2'.format(str(magic_ceil))
    for r,row in enumerate(magic_sheet['Y{0}:Y{1}'.format(str(template_floor), str(magic_ceil))],start=template_floor):
        for cell in row:
            cell.value = Y_formula.format(str(r),str(r))
    for r,row in enumerate(magic_sheet['AA{0}:AA{1}'.format(str(template_floor), str(magic_ceil))],start=template_floor):
        for cell in row:
            cell.value = AA_formula.format(str(r),str(r))

    # for r,row in enumerate(magic_sheet['AB{0}:AB{1}'.format(str(template_floor), str(magic_ceil))],start=template_floor):
    #     for cell in row:
    #         cell.value = AB_formula.format(str(r-1),str(r))
    #
    # for r,row in enumerate(magic_shecpet['Z{0}:Z{1}'.format(str(template_floor), str(magic_ceil))],start=template_floor):
    #     for cell in row:
    #         cell.value = Z_formula.format(str(r-1),str(r))

    for r,row in enumerate(magic_sheet['AB5:AB{0}'.format(str(magic_ceil))],start=5):
        for cell in row:
            cell.value = AB_formula.format(str(r-1),str(r))

    for r,row in enumerate(magic_sheet['Z5:Z{0}'.format(str(magic_ceil))],start=5):
        for cell in row:
            cell.value = Z_formula.format(str(r-1),str(r))

        # manual write


    magic_sheet.cell(row=magic_ceil,column=1).value = "Total Result"
    magic_sheet.cell(row=magic_ceil,column=2).value = sum(counts.values())
book.save(output_file)
