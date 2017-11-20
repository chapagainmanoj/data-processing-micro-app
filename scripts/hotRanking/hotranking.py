#!/usr/bin/env python

# NOTE: Here is issue with maxif funciton(@index 8 of company_formulae)

import re
import csv
import openpyxl
from sys import argv, maxsize
from openpyxl.utils.exceptions import IllegalCharacterError

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

# input_file = 'Drones - 3034 Companies.csv'
# another_input_file = 'Test - drones - target event report.csv'
# output_file = 'HotRanking.xlsx'
# template_file = 'template.xlsx'

input_file = argv[1]
another_input_file = argv[2]
output_file = argv[3]
template_file = argv[4]

company_formulae = [
    '=RANK(R{0},R:R)',\
    None,\
    '=VLOOKUP(B{0},\'Input - companies list\'!B:L,2,FALSE)',\
    '=VLOOKUP(B{0},\'Input - companies list\'!B:L,11,FALSE)',\
    '=VLOOKUP(B{0},\'Input - companies list\'!B:E,4,FALSE)',\
    '=SUMIFS(\'Input - target event report\'!H:H,\'Input - target event report\'!B:B,B{0},\'Input - target event report\'!D:D, "Private Investment")',\
    '=IF(I{0}<2, "N/A", (MAXIFS(\'Input - target event report\'!E:E,\'Input - target event report\'!B:B,B:B,\'Input - target event report\'!D:D,"Private Investment")-MINIFS(\'Input - target event report\'!E:E,\'Input - target event report\'!B:B,B:B,\'Input - target event report\'!D:D,"Private Investment"))/(I{0}-1))',\
    '=IF(MAXIFS(\'Input - target event report\'!E:E,\'Input - target event report\'!B:B,B:B,\'Input - target event report\'!D:D,"Private Investment") = 0, "N/A", TODAY() - MAXIFS(\'Input - target event report\'!E:E,\'Input - target event report\'!B:B,B:B,\'Input - target event report\'!D:D,"Private Investment"))',\
    '=COUNTIFS(\'Input - target event report\'!B:B,B{0},\'Input - target event report\'!D:D, "Private Investment")',\
    '=INDEX(\'Input - companies list\'!$1:$10000,MATCH(B{0},\'Input - companies list\'!B:B,0),MATCH("Flow",\'Input - companies list\'!$1:$1,0 ))',\
    '=INDEX(\'Input - companies list\'!$1:$10000,MATCH(B{0},\'Input - companies list\'!B:B,0),MATCH("Inter-Cluster Connectivity",\'Input - companies list\'!$1:$1,0 ))',\
    '=IFERROR(PERCENTRANK(F:F,F{0}),0)',\
    '=IFERROR(1 - PERCENTRANK(G:G,G{0}),0)',\
    '=IFERROR(1 - PERCENTRANK(H:H,H{0}),0)',\
    '=IFERROR(PERCENTRANK(I:I,I{0}),0)',\
    '=IFERROR(1 - PERCENTRANK(J:J,J{0}),0)',\
    '=IFERROR(PERCENTRANK(K:K,K{0}),0)',\
    '=L{0}*weight1+M{0}*weight2+N{0}*weight3+O{0}*weight4+P{0}*weight5+Q{0}*weight6'\
    ]
cluster_formulae = [
    '=RANK(M{0},M:M)',\
    None,\
    '=SUMIFS(\'Input - target event report\'!H:H,\'Input - target event report\'!L:L,B{0},\'Input - target event report\'!D:D, "Private Investment")',\
    '=IF(E{0}<2, "N/A", (MAXIFS(\'Input - target event report\'!E:E,\'Input - target event report\'!L:L,B:B,\'Input - target event report\'!D:D,"Private Investment")-MINIFS(\'Input - target event report\'!E:E,\'Input - target event report\'!L:L,B:B,\'Input - target event report\'!D:D,"Private Investment"))/(E{0}-1))',\
    '=COUNTIFS(\'Input - target event report\'!L:L,B{0},\'Input - target event report\'!D:D, "Private Investment")',\
    '=AVERAGEIF(INDEX(\'Input - companies list\'!$A:$BG, 0, MATCH("Clusters 0",\'Input - companies list\'!$A$1:$BG$1, 0)),\'Output - Cluster ranking\'!B{0},INDEX(\'Input - companies list\'!$A:$BG, 0, MATCH("Flow",\'Input - companies list\'!$A$1:$BG$1, 0)))',\
    '=AVERAGEIF(INDEX(\'Input - companies list\'!$A:$BG, 0, MATCH("Clusters 0",\'Input - companies list\'!$A$1:$BG$1, 0)),\'Output - Cluster ranking\'!B{0},INDEX(\'Input - companies list\'!$A:$BG, 0, MATCH("Inter-Cluster Connectivity",\'Input - companies list\'!$A$1:$BG$1, 0)))',\
    '=IFERROR(PERCENTRANK(C:C,C{0}),0)',\
    '=IFERROR(1 - PERCENTRANK(D:D,D{0}),0)',\
    '=IFERROR(PERCENTRANK(E:E,E{0}),0)',\
    '=IFERROR(1 - PERCENTRANK(F:F,F{0}),0)',\
    '=IFERROR(PERCENTRANK(G:G,G{0}),0)',\
    '=H{0}*weightA+I{0}*weightB+J{0}*weightC+K{0}*weightD+L{0}*weightE',\
]

# for OverflowError
decrement = True
while decrement:
    decrement = False
    try:
        csv.field_size_limit(maxsize)
    except OverflowError:
        maxsize = int(maxsize/10)
        decrement = True

clusters = set()
book = openpyxl.load_workbook(template_file)

ip_companies_list = book.get_sheet_by_name("Input - companies list")
ip_target_report = book.get_sheet_by_name("Input - target event report")
op_company_ranking = book.get_sheet_by_name("Output  - Company ranking")
op_cluster_ranking = book.get_sheet_by_name("Output - Cluster ranking")

with open(input_file, encoding='utf-8') as listFile:
    reader = csv.reader(listFile)
    for r, row in enumerate(reader,start=1):
        clusters |= set(list(map(str.strip,row[11].split('/'))))
        for c, cell in enumerate(row,start=1):
            try:
                ip_companies_list.cell(row=r,column=c).value = cell
            except IllegalCharacterError:
                cell = ILLEGAL_CHARACTERS_RE.sub('',cell)
                ip_companies_list.cell(row=r,column=c).value = cell

with open(another_input_file, encoding='utf-8') as targetFile:
    reader = csv.reader(targetFile)
    for r, row in enumerate(reader,start=1):
        # clusters |= set(list(map(str.strip,row['Cluster'.split(';'))))
        for c, cell in enumerate(row,start=1):
            try:
                ip_target_report.cell(row=r,column=c).value = cell
            except IllegalCharacterError:
                cell = ILLEGAL_CHARACTERS_RE.sub('',cell)
                ip_target_report.cell(row=r,column=c).value = cell

clusters.discard('Clusters 0')
clusters.discard('')

ip_companies_list_ceil = 2
ip_companies_list_floor = ip_companies_list.max_row
ip_target_report_ceil = 3
ip_target_report_floor = ip_companies_list_floor+(ip_target_report_ceil-ip_companies_list_ceil)
op_cluster_ranking_ceil = 3


for r in range(3,ip_companies_list_floor):
    for f in range(len(company_formulae)):
        if company_formulae[f]==None:
            pass
            # op_company_ranking.cell(row=r,column=(f+1)).value = company_formulae[f].format(str(r))
        else:
            op_company_ranking.cell(row=r,column=(f+1)).set_explicit_value(value=company_formulae[f].format(str(r)),data_type='f')

for src,dst in zip(ip_companies_list['B{0}:B{1}'.format(ip_companies_list_ceil,ip_companies_list_floor)],\
op_company_ranking['B{0}:B{1}'.format(ip_target_report_ceil,ip_target_report_floor)]):
    dst[0].value = src[0].value

for r in range(3,ip_companies_list_floor):
    for f in range(len(cluster_formulae)):
        if cluster_formulae[f]==None:
            pass
            # op_cluster_ranking.cell(row=r,column=(f+1)).value = cluster_formulae[f].format(str(r))
        else:
            op_cluster_ranking.cell(row=r,column=(f+1)).set_explicit_value(value=cluster_formulae[f].format(str(r)),data_type='f')

for src,dst in zip(clusters,op_cluster_ranking['B{0}:B{1}'.format(op_cluster_ranking_ceil,op_cluster_ranking_ceil+len(clusters))]):
    dst[0].value = src



book.save(output_file)
