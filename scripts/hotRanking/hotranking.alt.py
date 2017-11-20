#!/usr/bin/env python

import re
import csv
import openpyxl
from sys import argv, maxsize
from openpyxl.utils.exceptions import IllegalCharacterError

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

input_file = 'Drones - 3034 Companies.csv'
another_input_file = 'Test - drones - target event report.csv'
output_file = 'HotRanking.xlsx'
template_file = 'template.xlsx'

# input_file = argv[1]
# another_input_file = argv[2]
# output_file = argv[3]
# template_file = argv[4]

def copy_formulae(ws, r_index, t_index):
    row = []
    r,t= str(r_index),str(t_index)
    for x in (list(ws.rows)[r_index-1]):
        if x.value is not None:
            row.append(x.value.replace(r,t))
        else:
            row.append(x.value)
    return row

# for OverflowError
decrement = True
while decrement:
    decrement = False
    try:
        csv.field_size_limit(maxsize)
    except OverflowError:
        maxsize = int(maxsize/10)
        decrement = True

clusters = set()
book = openpyxl.load_workbook(template_file)

ip_companies_list = book.get_sheet_by_name("Input - companies list")
ip_target_report = book.get_sheet_by_name("Input - target event report")
op_company_ranking = book.get_sheet_by_name("Output  - Company ranking")
op_cluster_ranking = book.get_sheet_by_name("Output - Cluster ranking")

with open(input_file, encoding='utf-8') as listFile:
    reader = csv.reader(listFile)
    for r, row in enumerate(reader,start=1):
        clusters |= set(list(map(str.strip,row[11].split('/'))))
        for c, cell in enumerate(row,start=1):
            try:
                ip_companies_list.cell(row=r,column=c).value = cell
            except IllegalCharacterError:
                cell = ILLEGAL_CHARACTERS_RE.sub('',cell)
                ip_companies_list.cell(row=r,column=c).value = cell

with open(another_input_file, encoding='utf-8') as targetFile:
    reader = csv.reader(targetFile)
    for r, row in enumerate(reader,start=1):
        # clusters |= set(list(map(str.strip,row['Cluster'.split(';'))))
        for c, cell in enumerate(row,start=1):
            try:
                ip_target_report.cell(row=r,column=c).value = cell
            except IllegalCharacterError:
                cell = ILLEGAL_CHARACTERS_RE.sub('',cell)
                ip_target_report.cell(row=r,column=c).value = cell

ip_companies_list_ceil = 2
ip_companies_list_floor = ip_companies_list.max_row
ip_target_report_ceil = 3
ip_target_report_floor = ip_companies_list_floor+(ip_target_report_ceil-ip_companies_list_ceil)
op_cluster_ranking_ceil = 3

for r,row in enumerate(op_company_ranking['A500:R{}'.format(str(ip_companies_list_floor))],start=500):
    for cell, new in zip(row, copy_formulae(op_company_ranking,20,r)):
        cell.value = new

for src,dst in zip(ip_companies_list['B{0}:B{1}'.format(str(ip_companies_list_ceil),str(ip_companies_list_floor))], op_company_ranking['B{0}:B{1}'.format(str(ip_target_report_ceil),str(ip_target_report_floor))]):
    dst[0].value = src[0].value

clusters.discard('Clusters 0')
clusters.discard('')
for r,row in enumerate(op_cluster_ranking['A3:M{}'.format(str(op_cluster_ranking_ceil+len(clusters)))],start=3):
    for cell, new in zip(row, copy_formulae(op_cluster_ranking,14,r)):
        cell.value = new

for src,dst in zip(clusters,op_cluster_ranking['B{0}:B{1}'.format(str(op_cluster_ranking_ceil),str(op_cluster_ranking_ceil+len(clusters)))]):
    dst[0].value = src
book.save(output_file)
